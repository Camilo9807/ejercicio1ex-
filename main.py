import openpyxl

# 1. Entrada de datos y diccionario
estudiantes = {}  # Diccionario para almacenar los datos

for _ in range(3):  # Pide datos para 3 estudiantes
    nombre = input("Nombre del estudiante: ")
    nota = float(input("Nota del estudiante: "))  # Convierte la nota a float
    estudiantes[nombre] = nota

# 2. Creación del archivo Excel
libro = openpyxl.Workbook()  # Crea un nuevo libro de trabajo
hoja = libro.active  # Obtiene la hoja activa

# 3. Escritura de encabezados
hoja["A1"] = "Estudiante"
hoja["B1"] = "Nota"

# 4. Escritura de datos
fila = 2  # Comienza en la fila 2 para los datos
for nombre, nota in estudiantes.items():
    hoja.cell(row=fila, column=1, value=nombre)  # Escribe el nombre en la columna A
    hoja.cell(row=fila, column=2, value=nota)  # Escribe la nota en la columna B
    fila += 1  # Incrementa el número de fila

# 5. Guardar el archivo
libro.save("notas_estudiantes.xlsx")  # Guarda el archivo como "notas_estudiantes.xlsx"
print("Archivo Excel 'notas_estudiantes.xlsx' creado con éxito.")